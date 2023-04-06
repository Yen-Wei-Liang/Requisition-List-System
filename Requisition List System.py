"""

File: Requisition List System.py
Author: Yen, Wei-Liang　（KDD Lab 721）
Description: This program is used for recording the status of material collection by each student in the 『Embedded Real-Time Operating Systems for System-on-Chip』 course

"""


#===================================================================================
# Use of the library
#===================================================================================

import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from PIL import Image, ImageTk


#===================================================================================
# GUI window basic settings
#===================================================================================

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"
root = tk.Tk()                             
root.title("Embedded Real-Time Operating System for System-on-Chip rental material list system")    
root.geometry("960x540")                    
root.config(bg=background)                 
root.resizable(False,False)



#===================================================================================
# Function definition
#===================================================================================

def selection(variabl):

    """
    A function to record whether to receive various materials.

    Parameters:
    item0 : The first item.
    item1 : The second item.
          :      
    item24: Item 24.

    Returns:
    Whether to receive each weapon
    """

    global item0, item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13, item14, item15, item16, item17, item18, item19, item20, item21, item22, item23 
    value = variabl.get()
    
    if variabl == radio[0] and value == 1:
        item0 = "V"        
    elif variabl == radio[0] and value == 2:
        item0 = "X"
        
    if variabl == radio[1] and value == 1:
        item1 = "V"        
    elif variabl == radio[1] and value == 2:
        item1 = "X"

    if variabl == radio[2] and value == 1:
        item2 = "V"        
    elif variabl == radio[2] and value == 2:
        item2 = "X"

    if variabl == radio[3] and value == 1:
        item3 = "V"        
    elif variabl == radio[3] and value == 2:
        item3 = "X"        
        
    if variabl == radio[4] and value == 1:
        item4 = "V"        
    elif variabl == radio[4] and value == 2:
        item4 = "X"
        
    if variabl == radio[5] and value == 1:
        item5 = "V"        
    elif variabl == radio[5] and value == 2:
        item5 = "X"

    if variabl == radio[6] and value == 1:
        item6 = "V"        
    elif variabl == radio[6] and value == 2:
        item6 = "X"

    if variabl == radio[7] and value == 1:
        item7 = "V"        
    elif variabl == radio[7] and value == 2:
        item7 = "X"        
        
    if variabl == radio[8] and value == 1:
        item8 = "V"        
    elif variabl == radio[8] and value == 2:
        item8 = "X"
        
    if variabl == radio[9] and value == 1:
        item9 = "V"        
    elif variabl == radio[9] and value == 2:
        item9 = "X"

    if variabl == radio[10] and value == 1:
        item10 = "V"        
    elif variabl == radio[10] and value == 2:
        item10 = "X"

    if variabl == radio[11] and value == 1:
        item11 = "V"        
    elif variabl == radio[11] and value == 2:
        item11 = "X"        
        
    if variabl == radio[12] and value == 1:
        item12 = "V"        
    elif variabl == radio[12] and value == 2:
        item12 = "X"
        
    if variabl == radio[13] and value == 1:
        item13 = "V"        
    elif variabl == radio[13] and value == 2:
        item13 = "X"

    if variabl == radio[14] and value == 1:
        item14 = "V"        
    elif variabl == radio[14] and value == 2:
        item14 = "X"

    if variabl == radio[15] and value == 1:
        item15 = "V"        
    elif variabl == radio[15] and value == 2:
        item15 = "X"      
                
    if variabl == radio[16] and value == 1:
        item16 = "V"        
    elif variabl == radio[16] and value == 2:
        item16 = "X"
        
    if variabl == radio[17] and value == 1:
        item17 = "V"        
    elif variabl == radio[17] and value == 2:
        item17 = "X"

    if variabl == radio[18] and value == 1:
        item18 = "V"        
    elif variabl == radio[18] and value == 2:
        item18 = "X"

    if variabl == radio[19] and value == 1:
        item19 = "V"        
    elif variabl == radio[19] and value == 2:
        item19 = "X"        
        
    if variabl == radio[20] and value == 1:
        item20 = "V"        
    elif variabl == radio[20] and value == 2:
        item20 = "X"
        
    if variabl == radio[21] and value == 1:
        item21 = "V"        
    elif variabl == radio[21] and value == 2:
        item21 = "X"

    if variabl == radio[22] and value == 1:
        item22 = "V"        
    elif variabl == radio[22] and value == 2:
        item22 = "X"

    if variabl == radio[23] and value == 1:
        item23 = "V"        
    elif variabl == radio[23] and value == 2:
        item23 = "X"             


def Exit():
    """
    A function to close window.

    """
    root.destroy()



def Save():
    """
    This function checks whether the item has been received and includes an anti-fool feature. If the item has not been checked, a     reminder will be issued and the storage function cannot be used.
    
    Parameters:
    La : Lab.
    ID : Student ID.
       :      
    I23: Item 24.

    Remark:
    Foolproof will not be activated if no photos are uploaded.

    """

    La = Lab.get()
    ID = Student_ID.get()
    Na = Name.get()
    Em = Email.get()
    Li = Line.get()
    Te = Telephone.get()
    
    try:
        I0 = item0
        I1 = item1
        I2 = item2
        I3 = item3
        I4 = item4
        I5 = item5
        I6 = item6
        I7 = item7
        I8 = item8
        I9 = item9
        I10 = item10
        I11 = item11
        I12 = item12
        I13 = item13
        I14 = item14
        I15 = item15
        I16 = item16
        I17 = item17
        I18 = item18
        I19 = item19
        I20 = item20
        I21 = item21
        I22 = item22
        I23 = item23
    except:
        messagebox.showerror("error", "Select Claimed or Unclaimed")
    
    if La=="" or ID=="" or Na=="" or Em=="" or Li=="" or Te=="":
        messagebox.showerror("error", "Few Data is missing!")
    else:
        file = openpyxl.load_workbook('Embedded Real-Time Operating System for System-on-Chip rental material list system.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=La)
        sheet.cell(column=2, row=sheet.max_row, value=ID)
        sheet.cell(column=3, row=sheet.max_row, value=Na)
        sheet.cell(column=4, row=sheet.max_row, value=Em)
        sheet.cell(column=5, row=sheet.max_row, value=Te)
        sheet.cell(column=6, row=sheet.max_row, value=Li) 
        sheet.cell(column=7, row=sheet.max_row, value=I0)
        sheet.cell(column=8, row=sheet.max_row, value=I1)
        sheet.cell(column=9, row=sheet.max_row, value=I2)
        sheet.cell(column=10, row=sheet.max_row, value=I3)
        sheet.cell(column=11, row=sheet.max_row, value=I4)
        sheet.cell(column=12, row=sheet.max_row, value=I5)
        sheet.cell(column=13, row=sheet.max_row, value=I6)
        sheet.cell(column=14, row=sheet.max_row, value=I7)
        sheet.cell(column=15, row=sheet.max_row, value=I8)
        sheet.cell(column=16, row=sheet.max_row, value=I9)
        sheet.cell(column=17, row=sheet.max_row, value=I10)
        sheet.cell(column=18, row=sheet.max_row, value=I11)
        sheet.cell(column=19, row=sheet.max_row, value=I12)
        sheet.cell(column=20, row=sheet.max_row, value=I13)
        sheet.cell(column=21, row=sheet.max_row, value=I14)
        sheet.cell(column=22, row=sheet.max_row, value=I15)
        sheet.cell(column=23, row=sheet.max_row, value=I16)
        sheet.cell(column=24, row=sheet.max_row, value=I17)
        sheet.cell(column=25, row=sheet.max_row, value=I18)
        sheet.cell(column=26, row=sheet.max_row, value=I19)
        sheet.cell(column=27, row=sheet.max_row, value=I20)
        sheet.cell(column=28, row=sheet.max_row, value=I21)
        sheet.cell(column=29, row=sheet.max_row, value=I22)
        sheet.cell(column=30, row=sheet.max_row, value=I23)

        
        file.save('Embedded Real-Time Operating System for System-on-Chip rental material list system.xlsx')
        
        
#         try:
#             img.save("Student Images/"+Name.get()+".jpg")
#         except:
#             messagebox.showinfo("info", "Profile Picture is not available!!")
            
        
        messagebox.showinfo("showinfo", "Sucessfully Data Entered")
        Reset()


def Reset():
    """
    Function to initialize each input
    
    """
    global img
    Name.set('')
    Lab.set('')
    Student_ID.set('')
    Line.set('')
    Email.set('')
    Telephone.set('')
    for x in range(24):
        radio[x].set(None)
    SaveButton.config(state = 'normal')
    img = PhotoImage(file="kdd.png")
    lbl.config(image=img)
    lbl.image=img
    img=''






def showimage():
    """
    Function to upload personal photo

    """
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir = os.getcwd(),
                                          title="Select image file",
                                          filetype=(
                                              ("JPG File","*.jpg"),
                                              ("PNG File","*.png")
                                                    )
                                         )
    img = (Image.open(filename))
    resized_image = img.resize((150,150))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image = photo2)
    lbl.image = photo2



##################################The search function has not been successful#########################################
def search():
    text = search_var.get() 
    Reset() #清除空格
    SaveButton.config(state='disable')
    
    file = openpyxl.load_workbook("Embedded Real-Time Operating System for System-on-Chip rental material list system.xlsx")
    sheet = file.active
    
    for row in sheet.rows: 
        if row[1].value == str(text):
            name = row[0]
            reg_number = str(name)[15:-1]
            break
        else:
            messagebox.showerror("Invalid", "Invalid registeration number!!!")
            return

    if reg_number is not None:    
        x1 = sheet.cell(row=int(reg_number), column=1).value # lab
        x2 = sheet.cell(row=int(reg_number), column=2).value # ID
        x3 = sheet.cell(row=int(reg_number), column=3).value # Name
        x4 = sheet.cell(row=int(reg_number), column=4).value #E-mail
        x5 = sheet.cell(row=int(reg_number), column=5).value #Telephone
        x6 = sheet.cell(row=int(reg_number), column=6).value #Line
    
        Lab.set(x1)
        Student_ID.set(x2)
        Name.set(x3)
        Email.set(x4)
        Telephone.set(x5)   
        Line.set(x6)
######################################################################################################################



#===================================================================================
# Using excel as a database
#===================================================================================

file = pathlib.Path('Embedded Real-Time Operating System for System-on-Chip rental material list system.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Lab"
    sheet['B1'] = "Student ID"
    sheet['C1'] = "Name"
    sheet['D1'] = "E-mail"
    sheet['E1'] = "Telephone"
    sheet['F1'] = "Line"
    sheet['G1'] = "Arduino板*2"
    sheet['H1'] = "Arduino電源線*2"
    sheet['I1'] = "Arduino擴充版"
    sheet['J1'] = "充電電池*2"
    sheet['K1'] = "電池充電器*1"
    sheet['L1'] = "電池(2&4)盒各*1"
    sheet['M1'] = "馬達控制器*1"
    sheet['N1'] = "小車馬達*2"
    sheet['O1'] = "小車輔助輪*1"
    sheet['P1'] = "小車輪子*2"
    sheet['Q1'] = "小車板子*1"
    sheet['R1'] = "銅柱*1包"
    sheet['S1'] = "藍芽接收器*2"
    sheet['T1'] = "溫度感測器(DHT)*2"
    sheet['U1'] = "紅外線感測器*1"
    sheet['V1'] = "超音波感測器*1"
    sheet['W1'] = "LCD*1"
    sheet['X1'] = "杜邦線*1包"
    sheet['Y1'] = "樹梅派(3or4)*1"
    sheet['Z1'] = "記憶卡*1"
    sheet['AA1'] = "樹梅派相機*1"
    sheet['AB1'] = "樹梅派相機支架*1"
    sheet['AC1'] = "樹梅派(3or4)充電線*1"
    sheet['AD1'] = "Micro HDMI*1"
    file.save('Embedded Real-Time Operating System for System-on-Chip rental material list system.xlsx')


#===================================================================================
# Top UI frame
#===================================================================================

Label(root, text="Email:49937019@stust.edu.tw", width = 10, height=3, bg="#f0687c", anchor='e').pack(side = TOP,fill=X)
Label(root, text="Search for Student ID", width = 10, height=2, bg="#c36464", fg = '#fff', font ='arial 11 bold').pack(side = TOP,fill=X)

search_var = StringVar() 
search_entry = Entry(root, textvariable=search_var, width=10, bd=2, font="arial 13")
#search_entry.place(x=620, y=58)

srch = Button(root, text="Search", compound=LEFT, width=8, bg='#68ddfa', font='arial 13 bold', command=search)
#srch.place(x=800, y=58)  

#===================================================================================
# Left UI frame
#===================================================================================

f= Frame(root, bd=3,bg="black", width=150,height=150,relief=GROOVE)
f.place(x=15,y=130)
img = PhotoImage(file="kdd.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)



UploadButton =Button(root, text="Upload", width=14, height=2, font="arial 12 bold",bg="lightblue", command= showimage)
UploadButton.place(x=15, y=300)

SaveButton = Button(root, text="Save", width=14, height=2, font="arial 12 bold",bg="lightgreen", command= Save)
SaveButton.place(x=15, y=360)

Button(root, text="Reset", width=14, height=2, font="arial 12 bold",bg="lightpink", command= Reset).place(x=15, y=420)
Button(root, text="Exit", width=14, height=2, font="arial 12 bold",bg="grey", command= Exit).place(x=15, y=480)



#===================================================================================
# Upper right frame
#===================================================================================

# background
obj = LabelFrame(root, text="Student's Details", font= "arial 10",bd=2,width=700, bg=framebg, fg=framefg, height=130, relief=GROOVE).place(x=200,y=130)

#left label
Label(obj, text="Lab:", font="arial 11", bg = framebg, fg = framefg).place(x=250, y=160)
Label(obj, text="Student ID:", font="arial 11", bg = framebg, fg = framefg).place(x=250, y=190)
Label(obj, text="Name:", font="arial 11", bg = framebg , fg=framefg).place(x=250, y=220)

#right label
Label(obj, text="Line:", font="arial 11", bg = framebg , fg=framefg).place(x=580, y=160)
Label(obj, text="E-mail:", font="arial 11", bg = framebg , fg=framefg).place(x=580, y=190)
Label(obj, text="Telephone:", font="arial 11", bg = framebg , fg=framefg).place(x=580, y=220)


#left input box
Lab = StringVar()
Lab_entry = Entry(obj, textvariable = Lab, width=15, font="arial 11")
Lab_entry.place(x=350, y=160)


Student_ID = StringVar()
Student_ID_entry = Entry(obj, textvariable = Student_ID, width=15, font="arial 11")
Student_ID_entry.place(x=350, y=190)


Name = StringVar()
Name_entry = Entry(obj, textvariable = Name, width=15, font="arial 11")
Name_entry.place(x=350, y=220)


#right input box
Line = StringVar()
Line_entry = Entry(obj, textvariable = Line, width=15, font="arial 11")
Line_entry.place(x=710, y=160)

Email = StringVar()
Email_entry = Entry(obj, textvariable = Email, width=15, font="arial 11")
Email_entry.place(x=710, y=190)

Telephone = StringVar()
Telephone_entry = Entry(obj, textvariable = Telephone, width=15, font="arial 11")
Telephone_entry.place(x=710, y=220)

#===================================================================================
# Lower right frame
#===================================================================================


# background
obj2 = LabelFrame(root, text="Equipment Pickup List", font="arial 10",bd=2,width=700, bg=framebg, fg=framefg, height=265, relief=GROOVE).place(x=200,y=270)

# left label
Label(obj2, text="1、Arduino板*2", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=290)
Label(obj2, text="2、Arduino電源線*2", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=310)
Label(obj2, text="3、Arduino擴充版", font="arial 10", bg = framebg , fg=framefg).place(x=220, y=330)
Label(obj2, text="4、充電電池&盒*2", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=350)
Label(obj2, text="5、電池充電器*1", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=370)
Label(obj2, text="6、行動電源*1", font="arial 10", bg = framebg , fg=framefg).place(x=220, y=390)
Label(obj2, text="7、馬達控制器*1", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=410)
Label(obj2, text="8、小車馬達*2", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=430)
Label(obj2, text="9、小車輔助輪*1", font="arial 10", bg = framebg , fg=framefg).place(x=220, y=450)
Label(obj2, text="10、小車輪子*2", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=470)
Label(obj2, text="11、小車板子*1", font="arial 10", bg = framebg, fg = framefg).place(x=220, y=490)
Label(obj2, text="12、銅柱*1包", font="arial 10", bg = framebg , fg=framefg).place(x=220, y=510)

# Check button on the lift

radio = []
for num in range(24):
    radio.append('radio'+str(num))

for num in range(24):
    radio[num] = IntVar()
                                                                                                         
R1 = Radiobutton(obj2, text="Claimed", variable= radio[0], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[0])).place(x=350, y=290)
R2 = Radiobutton(obj2, text="UnClaimed", variable= radio[0], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[0])).place(x=440, y=290)
R3 = Radiobutton(obj2, text="Claimed", variable= radio[1], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[1])).place(x=350, y=310)
R4 = Radiobutton(obj2, text="Unclaimed", variable= radio[1], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[1])).place(x=440, y=310)
R5 = Radiobutton(obj2, text="Claimed", variable= radio[2], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[2])).place(x=350, y=330)
R6 = Radiobutton(obj2, text="Unclaimed", variable= radio[2], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[2])).place(x=440, y=330)
R7 = Radiobutton(obj2, text="Claimed", variable= radio[3], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[3])).place(x=350, y=350)
R8 = Radiobutton(obj2, text="Unclaimed", variable= radio[3], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[3])).place(x=440, y=350)
R9 = Radiobutton(obj2, text="Claimed", variable= radio[4], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[4])).place(x=350, y=370)
R10 = Radiobutton(obj2, text="Unclaimed", variable= radio[4], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[4])).place(x=440, y=370)
R11 = Radiobutton(obj2, text="Claimed", variable= radio[5], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[5])).place(x=350, y=390)
R12 = Radiobutton(obj2, text="Unclaimed", variable= radio[5], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[5])).place(x=440, y=390)
R13 = Radiobutton(obj2, text="Claimed", variable= radio[6], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[6])).place(x=350, y=410)
R14 = Radiobutton(obj2, text="Unclaimed", variable= radio[6], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[6])).place(x=440, y=410)
R15 = Radiobutton(obj2, text="Claimed", variable= radio[7], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[7])).place(x=350, y=430)
R16 = Radiobutton(obj2, text="Unclaimed", variable= radio[7], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[7])).place(x=440, y=430)
R17 = Radiobutton(obj2, text="Claimed", variable= radio[8], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[8])).place(x=350, y=450)
R18 = Radiobutton(obj2, text="Unclaimed", variable= radio[8], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[8])).place(x=440, y=450)
R19 = Radiobutton(obj2, text="Claimed", variable= radio[9], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[9])).place(x=350, y=470)
R20 = Radiobutton(obj2, text="Unclaimed", variable= radio[9], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[9])).place(x=440, y=470)
R21 = Radiobutton(obj2, text="Claimed", variable= radio[10], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[10])).place(x=350, y=490)
R22 = Radiobutton(obj2, text="Unclaimed", variable= radio[10], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[10])).place(x=440, y=490)
R23 = Radiobutton(obj2, text="Claimed", variable= radio[11], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[11])).place(x=350, y=510)
R24 = Radiobutton(obj2, text="Unclaimed", variable= radio[11], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[11])).place(x=440, y=510)



# right label
Label(obj2, text="13、藍芽接收器*2", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=290)
Label(obj2, text="14、溫度感測器(DHT)*2", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=310)
Label(obj2, text="15、紅外線感測器*1", font="arial 10", bg = framebg , fg=framefg).place(x=545, y=330)
Label(obj2, text="16、超音波感測器*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=350)
Label(obj2, text="17、LCD*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=370)
Label(obj2, text="18、杜邦線*1包", font="arial 10", bg = framebg , fg=framefg).place(x=545, y=390)
Label(obj2, text="19、樹梅派(3or4)*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=410)
Label(obj2, text="20、記憶卡*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=430)
Label(obj2, text="21、樹梅派相機*1", font="arial 10", bg = framebg , fg=framefg).place(x=545, y=450)
Label(obj2, text="22、樹梅派相機支架*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=470)
Label(obj2, text="23、樹梅派(3or4)充電線*1", font="arial 10", bg = framebg, fg = framefg).place(x=545, y=490)
Label(obj2, text="24、Micro HDMI*1", font="arial 10", bg = framebg , fg=framefg).place(x=545, y=510)

# Check button on the right
R25 = Radiobutton(obj2, text="claimed", variable= radio[12], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[12])).place(x=705, y=290)
R26 = Radiobutton(obj2, text="Unclaimed", variable= radio[12], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[12])).place(x=795, y=290)
R27 = Radiobutton(obj2, text="claimed", variable= radio[13], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[13])).place(x=705, y=310)
R28 = Radiobutton(obj2, text="Unclaimed", variable= radio[13], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[13])).place(x=795, y=310)
R29 = Radiobutton(obj2, text="claimed", variable= radio[14], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[14])).place(x=705, y=330)
R30 = Radiobutton(obj2, text="Unclaimed", variable= radio[14], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[14])).place(x=795, y=330)
R31 = Radiobutton(obj2, text="claimed", variable= radio[15], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[15])).place(x=705, y=350)
R32 = Radiobutton(obj2, text="Unclaimed", variable= radio[15], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[15])).place(x=795, y=350)
R33 = Radiobutton(obj2, text="claimed", variable= radio[16], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[16])).place(x=705, y=370)
R34 = Radiobutton(obj2, text="Unclaimed", variable= radio[16], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[16])).place(x=795, y=370)
R35 = Radiobutton(obj2, text="claimed", variable= radio[17], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[17])).place(x=705, y=390)
R36 = Radiobutton(obj2, text="Unclaimed", variable= radio[17], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[17])).place(x=795, y=390)
R37 = Radiobutton(obj2, text="claimed", variable= radio[18], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[18])).place(x=705, y=410)
R38 = Radiobutton(obj2, text="Unclaimed", variable= radio[18], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[18])).place(x=795, y=410)
R39 = Radiobutton(obj2, text="claimed", variable= radio[19], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[19])).place(x=705, y=430)
R40 = Radiobutton(obj2, text="Unclaimed", variable= radio[19], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[19])).place(x=795, y=430)
R41 = Radiobutton(obj2, text="claimed", variable= radio[20], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[20])).place(x=705, y=450)
R42 = Radiobutton(obj2, text="Unclaimed", variable= radio[20], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[20])).place(x=795, y=450)
R43 = Radiobutton(obj2, text="claimed", variable= radio[21], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[21])).place(x=705, y=470)
R44 = Radiobutton(obj2, text="Unclaimed", variable= radio[21], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[21])).place(x=795, y=470)
R45 = Radiobutton(obj2, text="claimed", variable= radio[22], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[22])).place(x=705, y=490)
R46 = Radiobutton(obj2, text="Unclaimed", variable= radio[22], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[22])).place(x=795, y=490)
R47 = Radiobutton(obj2, text="claimed", variable= radio[23], value=1, bg=framebg, fg=framefg, command = lambda: selection(radio[23])).place(x=705, y=510)
R48 = Radiobutton(obj2, text="Unclaimed", variable= radio[23], value=2, bg=framebg, fg=framefg, command = lambda: selection(radio[23])).place(x=795, y=510)



mainloop()
root.mainloop()  
